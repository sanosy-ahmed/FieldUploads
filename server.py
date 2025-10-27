import os
import traceback
from datetime import datetime
from fractions import Fraction as _Fraction

from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
from openpyxl import load_workbook, Workbook
from PIL import Image, ImageDraw, ImageFont

# ============== إعدادات أساسية ==============
# في Render لا يوجد تخزين دائم في الخطة المجانية، لذا نخزن مؤقتاً داخل الحاوية
# ثم نرفع النسخ النهائية إلى R2. سنستخدم مجلد "data" داخل المشروع كمسار عمل.
BASE_DIR = os.path.join(os.getcwd(), "data")
UPLOAD_FOLDER = BASE_DIR
IMAGES_FOLDER = os.path.join(UPLOAD_FOLDER, "images")
EXCEL_FILE = os.path.join(UPLOAD_FOLDER, "TaskLog.xlsx")

ENABLE_EXIF = True                 # محاولة كتابة EXIF GPS
STAMP_ON_SAVE = True               # طباعة نص معلومات على الصورة
MAX_UPLOAD_MB = 50                 # حد حجم الملف

os.makedirs(IMAGES_FOLDER, exist_ok=True)
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# ============== piexif (اختياري) ==============
try:
    import piexif
    PEX_OK = True
except Exception as e:
    print("[WARN] piexif not available:", e)
    PEX_OK = False

# ============== Cloudflare R2 عبر boto3 ==============
import boto3

R2_ACCOUNT_ID = os.getenv("R2_ACCOUNT_ID", "")
R2_ACCESS_KEY_ID = os.getenv("R2_ACCESS_KEY_ID", "")
R2_SECRET_ACCESS_KEY = os.getenv("R2_SECRET_ACCESS_KEY", "")
R2_BUCKET = os.getenv("R2_BUCKET", "")
# إن كان البكت Public فعّل هذا بالرابط العام (اختياري لكنه مريح)
R2_PUBLIC_BASE_URL = os.getenv("R2_PUBLIC_BASE_URL", "")  # مثال: https://fielduploads.<acc-id>.r2.cloudflarestorage.com

def get_r2_client():
    if not (R2_ACCOUNT_ID and R2_ACCESS_KEY_ID and R2_SECRET_ACCESS_KEY and R2_BUCKET):
        return None
    return boto3.client(
        "s3",
        endpoint_url=f"https://{R2_ACCOUNT_ID}.r2.cloudflarestorage.com",
        aws_access_key_id=R2_ACCESS_KEY_ID,
        aws_secret_access_key=R2_SECRET_ACCESS_KEY,
        region_name="auto",
    )

def r2_upload(local_path: str, key: str) -> bool:
    """يرفع ملفاً إلى R2 تحت المفتاح المعطى."""
    try:
        s3 = get_r2_client()
        if s3 is None:
            print("[INFO] R2 not configured; skipping upload")
            return False
        s3.upload_file(local_path, R2_BUCKET, key)
        return True
    except Exception as e:
        print("[WARN] R2 upload failed:", e)
        return False

def r2_public_url(key: str) -> str:
    """يبني رابطاً عاماً للملف إن كان لديك R2_PUBLIC_BASE_URL."""
    if R2_PUBLIC_BASE_URL:
        base = R2_PUBLIC_BASE_URL.rstrip('/')
        return f"{base}/{key.lstrip('/')}"
    return ""

# ============== Flask ==============
app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = MAX_UPLOAD_MB * 1024 * 1024
CORS(app)
LAST_ERR = None

# ============== أدوات EXIF ==============
def _rat(x: float):
    fr = _Fraction(x).limit_denominator()
    return (fr.numerator, fr.denominator)

def _deg_to_dms_rationals(deg_float: float):
    deg = int(abs(deg_float))
    minutes_float = (abs(deg_float) - deg) * 60
    minutes = int(minutes_float)
    seconds = (minutes_float - minutes) * 60
    return (_rat(deg), _rat(minutes), _rat(round(seconds, 2)))

def _sanitize_exif_dict(d):
    if not isinstance(d.get("0th"), dict): d["0th"] = {}
    if not isinstance(d.get("1st"), dict): d["1st"] = {}
    d["Exif"] = {}
    if "thumbnail" not in d: d["thumbnail"] = None
    if "GPS" not in d: d["GPS"] = {}
    return d

def write_gps_exif_jpeg_inplace(jpeg_path: str, lat_str: str, lon_str: str) -> bool:
    if not (ENABLE_EXIF and PEX_OK):
        return False
    try:
        lat = float(lat_str); lon = float(lon_str)
    except:
        print("[INFO] Skip EXIF: invalid lat/lon", lat_str, lon_str)
        return False

    lat_ref = "N" if lat >= 0 else "S"
    lon_ref = "E" if lon >= 0 else "W"
    lat_dms = _deg_to_dms_rationals(lat)
    lon_dms = _deg_to_dms_rationals(lon)

    try:
        try:
            exif_dict = piexif.load(jpeg_path)
        except Exception:
            exif_dict = {"0th": {}, "Exif": {}, "GPS": {}, "1st": {}, "thumbnail": None}
        exif_dict = _sanitize_exif_dict(exif_dict)

        gps = exif_dict["GPS"]
        gps[piexif.GPSIFD.GPSVersionID]    = (2, 3, 0, 0)
        gps[piexif.GPSIFD.GPSLatitudeRef]  = lat_ref.encode("ascii")
        gps[piexif.GPSIFD.GPSLatitude]     = lat_dms
        gps[piexif.GPSIFD.GPSLongitudeRef] = lon_ref.encode("ascii")
        gps[piexif.GPSIFD.GPSLongitude]    = lon_dms

        exif_bytes = piexif.dump(exif_dict)
        with Image.open(jpeg_path) as im:
            if im.mode != "RGB":
                im = im.convert("RGB")
            im.save(jpeg_path, "JPEG", quality=95, exif=exif_bytes)
        return True
    except Exception as e:
        print("[WARN] EXIF write failed:", e)
        return False

def ensure_jpeg(src_path: str) -> str:
    base, ext = os.path.splitext(src_path)
    ext = ext.lower()
    if ext in [".jpg", ".jpeg"]:
        return src_path
    dst_path = base + ".jpg"
    with Image.open(src_path) as im:
        if im.mode != "RGB":
            im = im.convert("RGB")
        im.save(dst_path, "JPEG", quality=95)
    try:
        os.remove(src_path)
    except:
        pass
    return dst_path

# ============== ختم نص على الصورة ==============
def stamp_text_on_image(img_path, lines, margin=12, line_spacing=10, scale=5):
    """
    يطبع سطور معلومات باللون الأبيض أسفل يسار الصورة (بدون خلفية).
    يستخدم خط PIL الافتراضي (قد لا يعرض العربية مثاليًا لكن الأرقام/لاتيني ممتازة).
    scale لتكبير الخط (5≈ كبير).
    """
    try:
        with Image.open(img_path).convert("RGB") as im:
            draw = ImageDraw.Draw(im)
            # لا نعتمد على TTF. نستخدم خط PIL الافتراضي ونكرر الرسم لتكبير بصري.
            font = ImageFont.load_default()
            W, H = im.size

            # احسب إجمالي الارتفاع
            total_h = 0
            heights = []
            for line in lines:
                bbox = draw.textbbox((0, 0), line, font=font)
                h = (bbox[3] - bbox[1]) * scale
                heights.append(h)
                total_h += h + line_spacing
            total_h -= line_spacing

            x = margin
            y = H - margin - total_h

            # ارسم كل سطر مكبّرًا بتكرار الطباعة (scale مرات عموديًا وأفقيًا)
            for idx, line in enumerate(lines):
                # طباعة مكبرة: شبكة scale x scale
                for dx in range(scale):
                    for dy in range(scale):
                        draw.text((x + dx, y + dy), line, fill=(255, 255, 255), font=font)
                y += heights[idx] + line_spacing

            im.save(img_path, "JPEG", quality=95)
            return True
    except Exception as e:
        print("[WARN] stamping failed:", e)
        return False

# ============== Excel ==============
def ensure_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "TaskLog"
        ws.append(["رقم المهمة", "رقم المحطة", "ملاحظات", "اسم الصورة",
                   "Latitude", "Longitude", "Timestamp", "Image URL"])
        wb.save(EXCEL_FILE)

def append_record(task_id, station_id, note, img_filename, lat, lon, img_url):
    ensure_excel()
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb["TaskLog"]
        ws.append([task_id, station_id, note, img_filename, lat, lon, ts, img_url])
        link_cell = ws.cell(row=ws.max_row, column=8)
        link_cell.hyperlink = img_url
        link_cell.style = "Hyperlink"
        try:
            wb.save(EXCEL_FILE)
        except PermissionError:
            tmp = os.path.join(UPLOAD_FOLDER, f"TaskLog_tmp_{datetime.now().strftime('%H%M%S')}.xlsx")
            wb.save(tmp)
            os.replace(tmp, EXCEL_FILE)
    except Exception:
        with open(os.path.join(UPLOAD_FOLDER, "TaskLog_fallback.csv"), "a", encoding="utf-8") as f:
            f.write(f"{task_id},{station_id},{note},{img_filename},{lat},{lon},{ts},{img_url}\n")

# ============== Routes ==============
@app.get("/ping")
def ping():
    return "pong"

@app.get("/")
def root():
    return "<h3>FieldUploads server running ✅</h3><a href='/gallery'>Gallery</a>"

@app.get("/images/<path:fname>")
def get_image(fname):
    return send_from_directory(IMAGES_FOLDER, fname, as_attachment=False)

@app.get("/gallery")
def gallery():
    files = [f for f in os.listdir(IMAGES_FOLDER)
             if f.lower().endswith((".jpg", ".jpeg", ".png", ".webp"))]
    files.sort(reverse=True)
    cards = "".join(
        f"<div style='margin:8px'><img src='/images/{f}' style='max-width:240px;display:block'><small>{f}</small></div>"
        for f in files[:100]
    )
    return f"<html><body style='font-family:sans-serif;padding:12px'><h3>Gallery</h3><div style='display:flex;flex-wrap:wrap'>{cards}</div></body></html>"

@app.get("/debug_last")
def debug_last():
    global LAST_ERR
    return jsonify({"ok": LAST_ERR is None, "last_error": LAST_ERR}), 200 if LAST_ERR is None else 500

@app.post("/upload")
def upload():
    global LAST_ERR
    LAST_ERR = None
    try:
        if 'image' not in request.files:
            return jsonify({"ok": False, "error": "no image part"}), 400

        image = request.files['image']
        task_id    = (request.form.get('task_id')    or '').strip()
        station_id = (request.form.get('station_id') or '').strip()
        note       = (request.form.get('note')       or '').strip()
        latitude   = (request.form.get('latitude')   or '').strip()
        longitude  = (request.form.get('longitude')  or '').strip()

        if not task_id:
            return jsonify({"ok": False, "error": "task_id required"}), 400
        if image.filename == '':
            return jsonify({"ok": False, "error": "empty filename"}), 400

        # احفظ كما وصل
        raw_name = f"{task_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}_{image.filename}"
        raw_path = os.path.join(IMAGES_FOLDER, raw_name)
        image.save(raw_path)

        # ضمن JPG
        final_path = ensure_jpeg(raw_path)
        final_name = os.path.basename(final_path)

        # حاول رفع الصورة إلى R2 أولاً
        r2_key_img = f"images/{final_name}"
        r2_ok = r2_upload(final_path, r2_key_img)

        # حدد رابط الصورة المعلن
        if r2_ok and R2_PUBLIC_BASE_URL:
            img_url = r2_public_url(r2_key_img)
        else:
            base = request.host_url.rstrip('/')
            img_url = f"{base}/images/{final_name}"

        # سجّل في Excel (مع الرابط)
        append_record(task_id, station_id, note, final_name, latitude, longitude, img_url)

        # حاول كتابة EXIF GPS (لا يوقف العملية لو فشل)
        exif_ok = write_gps_exif_jpeg_inplace(final_path, latitude, longitude)

        # طباعة نص معلومات على الصورة (اختياري)
        stamp_ok = False
        if STAMP_ON_SAVE:
            ts_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            lines = [
                f"Task: {task_id}  |  Station: {station_id}",
                f"Lat: {latitude}  |  Lon: {longitude}",
                f"Time: {ts_str}",
                f"File: {final_name}",
            ]
            stamp_ok = stamp_text_on_image(final_path, lines, scale=5)

        # ارفع Excel أيضاً إلى R2 لضمان نسخة دائمة
        r2_upload(EXCEL_FILE, "TaskLog.xlsx")

        return jsonify({
            "ok": True,
            "saved": final_name,
            "url": img_url,
            "exif_gps_written": bool(exif_ok and ENABLE_EXIF and PEX_OK),
            "stamped": stamp_ok,
            "r2_upload": r2_ok
        }), 200

    except Exception as e:
        LAST_ERR = f"{type(e).__name__}: {e}"
        traceback.print_exc()
        return jsonify({"ok": False, "error": f"server-error: {type(e).__name__}: {e}"}), 500

if __name__ == "__main__":
    port = int(os.environ.get("PORT", "5000"))
    print(f"[INFO] Server starting on 0.0.0.0:{port}")
    app.run(host="0.0.0.0", port=port, debug=False)
